# Q8. Write a Python program that:
# • Reads a paragraph from a text file.
# • Removes punctuation and numbers.
# • Extracts all email IDs and hashtags using Regular Expressions (re module).
# • Saves the cleaned text to an Excel file using Openpyxl.



import re
from openpyxl import Workbook

# --- Step 1: Read paragraph from a text file ---
input_file = "sample_text.txt"  # replace with your file name

try:
    with open(input_file, "r", encoding="utf-8") as f:
        text = f.read()
        print("✅ File read successfully!\n")
except FileNotFoundError:
    print("❌ File not found! Please make sure sample_text.txt is in the same folder.")
    exit()

# --- Step 2: Extract email IDs and hashtags using regex ---
emails = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", text)
hashtags = re.findall(r"#\w+", text)

# --- Step 3: Remove punctuation and numbers from text ---
cleaned_text = re.sub(r"[^A-Za-z\s]", "", text)  # keep only letters and spaces
cleaned_text = re.sub(r"\s+", " ", cleaned_text)  # remove extra spaces

# --- Step 4: Display results in console ---
print("📧 Emails found:", emails)
print("🏷️ Hashtags found:", hashtags)
print("\n🧹 Cleaned text:\n", cleaned_text)

# --- Step 5: Save cleaned text, emails, and hashtags to Excel ---
wb = Workbook()
ws = wb.active
ws.title = "Text_Cleaning_Result"

# Write headings
ws["A1"] = "Cleaned Text"
ws["B1"] = "Emails"
ws["C1"] = "Hashtags"

# Fill in data
ws["A2"] = cleaned_text
ws["B2"] = ", ".join(emails) if emails else "No Emails Found"
ws["C2"] = ", ".join(hashtags) if hashtags else "No Hashtags Found"

# Save to Excel file
output_file = "result_q8.xlsx"
wb.save(output_file)
print(f"\n💾 Data saved successfully to {output_file}")
